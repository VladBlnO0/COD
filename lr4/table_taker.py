import openpyxl

wb = openpyxl.load_workbook("bank_cool_data.xlsx")

ws = wb["clients"]

print(
    "Total number of rows: "
    + str(ws.max_row)
    + ". And total number of columns: "
    + str(ws.max_column)
)

clients: list[str] = [str(ws.cell(row=i, column=1).value) for i in range(2, 22)]


def get_value(cell_value):
    try:
        return float(cell_value)
    except ValueError:
        return 0.0


balances: list[float] = [
    get_value(ws.cell(row=i, column=2).value) for i in range(3, 22)
]

sorted_cb = sorted(list(zip(clients, balances)), reverse=True, key=lambda item: item[1])

print(f"{'Client':<10} {' ':>5} {'Balance':>10}")
for cli, bal in zip(clients, balances):
    print(f"{cli:<10} {'-':>5} {bal:>10}")

ws.cell(row=1, column=3, value=sorted_cb[0][0])
ws.cell(row=2, column=3, value=sorted_cb[0][1])
wb.save("bank_cool_data.xlsx")
print("The richest client is: " + str((ws["C1"]).value) + " - " + str(ws["C2"].value))
